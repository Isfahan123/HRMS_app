"""
Excel and CSV file parsing utilities without pandas.

This module provides file parsing functionality using openpyxl for Excel files
and the built-in csv module for CSV files. This ensures compatibility with
cPanel/shared hosting environments that cannot compile numpy/pandas.
"""
import csv
from typing import List, Dict, Any, Optional, Tuple
from openpyxl import load_workbook


def normalize_column_name(name: str) -> str:
    """
    Normalize column name by stripping whitespace and replacing special characters.
    
    Args:
        name: Original column name
        
    Returns:
        Normalized column name (lowercase, underscores, no special chars)
    """
    return (
        str(name).strip().lower()
        .replace(" ", "_")
        .replace("(", "")
        .replace(")", "")
        .replace("'", "")
        .replace(",", "")
    )


def read_csv_as_dicts(file_path: str) -> List[Dict[str, Any]]:
    """
    Read a CSV file and return a list of dictionaries.
    
    Args:
        file_path: Path to the CSV file
        
    Returns:
        List of dictionaries where keys are normalized column names
    """
    rows = []
    with open(file_path, 'r', newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        headers = None
        for i, row in enumerate(reader):
            if i == 0:
                # First row is headers
                headers = [normalize_column_name(h) for h in row]
                continue
            if headers and row:
                row_dict = {}
                for j, cell in enumerate(row):
                    if j < len(headers):
                        row_dict[headers[j]] = cell.strip() if cell else ""
                rows.append(row_dict)
    return rows


def read_excel_as_dicts(file_path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    Read an Excel file and return a list of dictionaries.
    
    Args:
        file_path: Path to the Excel file (.xlsx or .xls)
        sheet_name: Optional sheet name. Uses first sheet if not specified.
        
    Returns:
        List of dictionaries where keys are normalized column names
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    
    rows = []
    headers = None
    
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            # First row is headers
            headers = [normalize_column_name(h) if h else f"col_{j}" for j, h in enumerate(row)]
            continue
        if headers and row:
            # Skip empty rows
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            row_dict = {}
            for j, cell in enumerate(row):
                if j < len(headers):
                    # Convert cell value to string, handle None
                    if cell is None:
                        row_dict[headers[j]] = ""
                    else:
                        row_dict[headers[j]] = cell
            rows.append(row_dict)
    
    wb.close()
    return rows


def find_column_by_names(row_dict: Dict[str, Any], possible_names: List[str]) -> Optional[str]:
    """
    Find the first matching column name from a list of possibilities.
    
    Args:
        row_dict: Dictionary with column names as keys
        possible_names: List of possible column names to try
        
    Returns:
        The first matching column name found, or None
    """
    for name in possible_names:
        normalized = normalize_column_name(name)
        if normalized in row_dict:
            return normalized
    return None


def get_column_names(data: List[Dict[str, Any]]) -> List[str]:
    """
    Get all unique column names from a list of row dictionaries.
    
    Args:
        data: List of row dictionaries
        
    Returns:
        List of unique column names
    """
    if not data:
        return []
    return list(data[0].keys())


def parse_wage_range(wage_str: str) -> Tuple[float, float]:
    """
    Parse a wage range string into (min, max) values.
    
    Handles formats like:
    - "0.00 - 30.00"
    - "6000.01 and above"
    - "RM 100 – 200"
    
    Args:
        wage_str: String representing wage range
        
    Returns:
        Tuple of (wage_min, wage_max) as floats
    """
    wage_str = str(wage_str).strip()
    
    # Handle "and above" case
    if "and above" in wage_str.lower():
        # Extract the number before "and above"
        parts = wage_str.lower().split("and above")[0].strip()
        # Remove any RM prefix and get the number
        parts = parts.replace("rm", "").replace(",", "").strip()
        try:
            wage_min = float(parts.split()[-1])
        except (ValueError, IndexError):
            wage_min = float(parts)
        return (wage_min, 999999.99)
    
    # Clean up the string
    wage_str = wage_str.replace("RM", "").replace("rm", "").replace(",", "")
    
    # Handle different separators: - – (en dash) — (em dash)
    for sep in [" - ", " – ", " — ", "-", "–", "—"]:
        if sep in wage_str:
            parts = wage_str.split(sep)
            if len(parts) == 2:
                try:
                    wage_min = float(parts[0].strip())
                    wage_max = float(parts[1].strip())
                    return (wage_min, wage_max)
                except ValueError:
                    continue
    
    # If no range found, try to parse as single value
    try:
        val = float(wage_str.strip())
        return (val, val)
    except ValueError:
        raise ValueError(f"Could not parse wage range: {wage_str}")
